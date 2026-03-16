"""
Excel export — stap-voor-stap berekening als narekenbaar Excel-bestand.

Genereert een workbook met 5 sheets:
  1. Scenario — geselecteerde strategie en parameters
  2. Berekening — financiële stappen met live Excel-formules
  3. Kwartaaloverzicht — aggregatie per Q1-Q4
  4. Uurdata — volledig DataFrame met alle berekende kolommen
  5. Uitleg — methodologie en kolomdefinities

Geen Streamlit-afhankelijkheden.
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config as cfg

# ---------------------------------------------------------------------------
# Styling constanten
# ---------------------------------------------------------------------------

_GOLD_HEX = "FAB517"
_GREY_HEX = "9E9E9E"

_HEADER_FILL = PatternFill(start_color=_GOLD_HEX, end_color=_GOLD_HEX, fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="000000")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="000000")
_LABEL_FONT = Font(name="Calibri", bold=True, size=10)
_DATA_FONT = Font(name="Calibri", size=10)
_GREY_FONT = Font(name="Calibri", italic=True, size=9, color=_GREY_HEX)
_SECTION_FONT = Font(name="Calibri", bold=True, size=11, color=_GOLD_HEX)

# Uurdata kolommen: (df_kolom, excel_header, number_format | None)
_UURDATA_COLS = [
    ("Date", "Datum/tijd", "DD-MM-YYYY HH:MM"),
    ("Quarter", "Kwartaal", "0"),
    ("is_peak", "Piekuur (1=ja)", "0"),
    ("Active_Profile_MW", "Profiel (MW)", "#,##0.00"),
    ("Price_Base", "Base prijs (€/MWh)", "#,##0.00"),
    ("Price_Peak", "Peak prijs (€/MWh)", "#,##0.00"),
    ("EPEX_EUR_MWh", "EPEX spot (€/MWh)", "#,##0.00"),
    ("Hedge_Base_MW", "Base inkoop (MW)", "#,##0.00"),
    ("Hedge_Peak_MW", "Peak inkoop (MW)", "#,##0.00"),
    ("Current_Hedge_MW", "Totaal inkoop (MW)", "#,##0.00"),
    ("Profile_MWh", "Profiel (MWh)", "#,##0.000"),
    ("Hedge_MWh", "Inkoop (MWh)", "#,##0.000"),
    ("Over_Hedge_MWh", "Overschot (MWh)", "#,##0.000"),
    ("Under_Hedge_MWh", "Tekort (MWh)", "#,##0.000"),
    ("Used_Hedge_MWh_Abs", "Benut (MWh)", "#,##0.000"),
    ("Cost_Hedge_Base_EUR", "Kosten base (€)", '#,##0.00'),
    ("Cost_Hedge_Peak_EUR", "Kosten peak (€)", '#,##0.00'),
    ("Cost_Hedge_Total_EUR", "Kosten inkoop (€)", '#,##0.00'),
    ("Cost_Buy_EUR", "Spot inkoop (€)", '#,##0.00'),
    ("Rev_Sell_EUR", "Spot verkoop (€)", '#,##0.00'),
    ("Net_Spot_EUR", "Netto spot (€)", '#,##0.00'),
]


def _col_letter(col_name: str) -> str:
    """Excel-kolomletter voor een DataFrame-kolomnaam in de Uurdata sheet."""
    for i, (name, _, _) in enumerate(_UURDATA_COLS):
        if name == col_name:
            return get_column_letter(i + 1)
    raise ValueError(f"Kolom {col_name!r} niet gevonden in _UURDATA_COLS")


def _style_header_row(ws, row: int, n_cols: int) -> None:
    """Pas goud-achtergrond + bold font toe op een header-rij."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    """Stel kolom-breedtes in. Keys zijn kolomletters."""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1: Scenario
# ═══════════════════════════════════════════════════════════════════════════

def _build_scenario_sheet(
    wb: Workbook,
    category: str,
    optimization_key: str,
    strategy_period: str,
    base_mw: float,
    peak_add_mw: float,
    price_base: float,
    price_peak: float,
) -> None:
    ws = wb.active
    ws.title = "Scenario"

    ws["A1"] = "Censo Energy Hedge — Scenario-overzicht"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:C1")

    ws["A3"] = "Export datum"
    ws["B3"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A3"].font = _LABEL_FONT

    opt = next((o for o in cfg.OPTIMIZATIONS if o["key"] == optimization_key), None)

    info_rows = [
        ("Categorie", category),
        ("", ""),
        ("Strategie", opt["label"] if opt else optimization_key),
        ("Toelichting", opt["desc"] if opt else ""),
        ("", ""),
        ("Contractperiode", strategy_period),
        ("Base prijs (€/MWh)", f"{price_base:.2f}"),
        ("Peak prijs (€/MWh)", f"{price_peak:.2f}"),
        ("", ""),
        ("Resultaat optimalisatie", ""),
        ("Base MW", f"{base_mw:.1f}"),
        ("Peak Add MW", f"{peak_add_mw:.1f}"),
        ("Totaal Peak MW", f"{base_mw + peak_add_mw:.1f}"),
    ]

    for i, (label, value) in enumerate(info_rows, start=5):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)
        cell_a.font = _LABEL_FONT if label else _DATA_FONT
        cell_b.font = _DATA_FONT

    _set_col_widths(ws, {"A": 30, "B": 60})


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2: Berekening (met live Excel-formules)
# ═══════════════════════════════════════════════════════════════════════════

def _build_berekening_sheet(wb: Workbook, n_rows: int) -> None:
    ws = wb.create_sheet("Berekening")

    ws["A1"] = "Stap-voor-stap berekening"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A2"] = (
        "Alle waarden worden berekend met Excel-formules die verwijzen naar "
        "het tabblad 'Uurdata'. Klik op een cel in kolom C om de formule te zien."
    )
    ws["A2"].font = _GREY_FONT
    ws.merge_cells("A2:E2")

    # Headers
    headers = ["Stap", "Beschrijving", "Waarde", "Eenheid", "Toelichting"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=4, column=j, value=h)
    _style_header_row(ws, 4, len(headers))

    # Kolomletters voor formules
    last = n_rows + 1  # data begint op rij 2 in Uurdata
    prof = _col_letter("Profile_MWh")
    hedge = _col_letter("Hedge_MWh")
    over = _col_letter("Over_Hedge_MWh")
    under = _col_letter("Under_Hedge_MWh")
    used = _col_letter("Used_Hedge_MWh_Abs")
    cost_h = _col_letter("Cost_Hedge_Total_EUR")
    cost_b = _col_letter("Cost_Buy_EUR")
    rev_s = _col_letter("Rev_Sell_EUR")

    # Stappen: (excel_row, stap_nr, beschrijving, formula, eenheid, toelichting, fmt)
    steps = [
        # ── VOLUME ──
        (5, "", "VOLUME", None, "", "", None),
        (6, 1, "Totaal profiel (verbruik)",
         f"=SUM(Uurdata!{prof}2:{prof}{last})", "MWh",
         f"= Profiel (MW) × {cfg.MWH_FACTOR} per 15-min interval",
         "#,##0.0"),
        (7, 2, "Totaal ingekocht via blokken",
         f"=SUM(Uurdata!{hedge}2:{hedge}{last})", "MWh",
         "= (Base + Peak) × 0.25 per interval",
         "#,##0.0"),
        (8, 3, "Tekort (spot inkoop nodig)",
         f"=SUM(Uurdata!{under}2:{under}{last})", "MWh",
         "= MAX(0, Profiel − Inkoop) per interval",
         "#,##0.0"),
        (9, 4, "Overschot (spot verkoop)",
         f"=SUM(Uurdata!{over}2:{over}{last})", "MWh",
         "= MAX(0, Inkoop − Profiel) per interval",
         "#,##0.0"),
        (10, 5, "Benut uit inkoop",
         f"=SUM(Uurdata!{used}2:{used}{last})", "MWh",
         "= MIN(|Profiel|, |Inkoop|) als zelfde richting",
         "#,##0.0"),
        (11, 6, "Hedge-efficiëntie",
         "=IF(ABS(C6)>0,C10/ABS(C6)*100,0)", "%",
         "= Benut / |Profiel| × 100",
         "#,##0.0"),

        # ── FINANCIEEL ──
        (13, "", "FINANCIEEL", None, "", "", None),
        (14, 7, "Kosten vaste inkoop (ENDEX)",
         f"=SUM(Uurdata!{cost_h}2:{cost_h}{last})", "€",
         "= Base×Baseprijs + Peak×Peakprijs per interval",
         '€ #,##0'),
        (15, 8, "Kosten spot inkoop",
         f"=SUM(Uurdata!{cost_b}2:{cost_b}{last})", "€",
         "= Tekort × EPEX spotprijs",
         '€ #,##0'),
        (16, 9, "Opbrengst spot verkoop",
         f"=SUM(Uurdata!{rev_s}2:{rev_s}{last})", "€",
         "= Overschot × EPEX spotprijs",
         '€ #,##0'),
        (17, 10, "Netto spotresultaat",
         "=C16-C15", "€",
         "= Opbrengst − Inkoop",
         '€ #,##0'),

        # ── TOTAAL ──
        (19, "", "TOTAAL", None, "", "", None),
        (20, 11, "Totale energiekosten",
         "=C14-C17", "€",
         "= Vaste inkoop − Netto spotresultaat",
         '€ #,##0'),
        (21, 12, "Kostprijs per MWh",
         "=IF(ABS(C6)>0,C20/ABS(C6),0)", "€/MWh",
         "= Totale kosten / |Profiel|",
         '€ #,##0.00'),
    ]

    for row, stap, beschr, formula, eenh, toel, fmt in steps:
        # Sectieheaders
        if formula is None and stap == "":
            ws.cell(row=row, column=2, value=beschr).font = _SECTION_FONT
            continue

        ws.cell(row=row, column=1, value=stap).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=beschr).font = _DATA_FONT
        cell_c = ws.cell(row=row, column=3, value=formula)
        cell_c.font = Font(name="Calibri", bold=True, size=11)
        if fmt:
            cell_c.number_format = fmt
        ws.cell(row=row, column=4, value=eenh).font = _DATA_FONT
        ws.cell(row=row, column=5, value=toel).font = _GREY_FONT

    _set_col_widths(ws, {"A": 8, "B": 35, "C": 18, "D": 10, "E": 50})


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 3: Kwartaaloverzicht
# ═══════════════════════════════════════════════════════════════════════════

def _build_quarterly_sheet(wb: Workbook, q_stats: pd.DataFrame) -> None:
    ws = wb.create_sheet("Kwartaaloverzicht")

    ws["A1"] = "Kwartaaloverzicht"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(q_stats.columns) + 1)}1")

    # Headers
    ws.cell(row=3, column=1, value="Kwartaal")
    for j, col_name in enumerate(q_stats.columns, 2):
        ws.cell(row=3, column=j, value=col_name)
    _style_header_row(ws, 3, len(q_stats.columns) + 1)

    # Data
    fmt_map = {
        "Volume (MWh)": "#,##0",
        "Afgedekt (%)": "#,##0.0",
        "Verkocht (MWh)": "#,##0",
        "Ingekocht (MWh)": "#,##0",
        "Spot resultaat (€)": '€ #,##0',
        "Totale kosten (€)": '€ #,##0',
        "Jouw prijs (€/MWh)": '€ #,##0.00',
    }

    for i, (q_idx, row_data) in enumerate(q_stats.iterrows(), 4):
        ws.cell(row=i, column=1, value=f"Q{q_idx}").font = _LABEL_FONT
        for j, col_name in enumerate(q_stats.columns, 2):
            cell = ws.cell(row=i, column=j, value=row_data[col_name])
            cell.font = _DATA_FONT
            if col_name in fmt_map:
                cell.number_format = fmt_map[col_name]

    # Kolom-breedtes
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(q_stats.columns) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 18


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 4: Uurdata
# ═══════════════════════════════════════════════════════════════════════════

def _build_uurdata_sheet(wb: Workbook, df: pd.DataFrame) -> int:
    """Schrijf het volledige DataFrame naar de Uurdata sheet.

    Returns het aantal datarijen (excl. header).
    """
    ws = wb.create_sheet("Uurdata")

    # Filter beschikbare kolommen
    available = [(col, label, fmt) for col, label, fmt in _UURDATA_COLS if col in df.columns]
    col_names = [col for col, _, _ in available]
    col_labels = [label for _, label, _ in available]
    col_fmts = [fmt for _, _, fmt in available]

    # Headers
    ws.append(col_labels)
    _style_header_row(ws, 1, len(col_labels))

    # Data — converteer Date-kolom naar Python datetimes
    export_df = df[col_names].copy()
    if "Date" in export_df.columns:
        export_df["Date"] = export_df["Date"].dt.to_pydatetime()

    for row_data in export_df.values.tolist():
        ws.append(row_data)

    # Number formatting per kolom
    n_rows = len(export_df)
    for col_idx, fmt in enumerate(col_fmts, 1):
        if fmt:
            for row_idx in range(2, n_rows + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # Kolom-breedtes
    for col_idx in range(1, len(col_labels) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze header
    ws.freeze_panes = "A2"

    return n_rows


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 5: Uitleg
# ═══════════════════════════════════════════════════════════════════════════

def _build_uitleg_sheet(
    wb: Workbook,
    optimization_key: str,
) -> None:
    ws = wb.create_sheet("Uitleg")

    ws["A1"] = "Uitleg methodologie"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:B1")

    opt = next((o for o in cfg.OPTIMIZATIONS if o["key"] == optimization_key), None)

    sections = [
        ("Wat is een hedge?",
         "Een hedge is het vooraf inkopen van energie via vaste ENDEX-contracten "
         "(blokken) om je te beschermen tegen prijsschommelingen op de spotmarkt (EPEX). "
         "Je koopt een vast aantal MW in voor een bepaalde periode."),

        ("Base en Peak blokken",
         f"Base: 24/7 constante inkoop — alle uren van de week.\n"
         f"Peak: extra inkoop tijdens piekuren — Ma t/m Vr, "
         f"{cfg.PEAK_HOUR_START}:00 – {cfg.PEAK_HOUR_END}:00.\n"
         f"Totaal inkoop per uur = Base MW + (Peak MW × is_piekuur)."),

        ("Berekening per 15-min interval",
         f"Elke rij in het Uurdata-tabblad is een 15-minuten interval.\n"
         f"MWh-conversie: MW × {cfg.MWH_FACTOR} (kwartier = 0.25 uur).\n\n"
         f"Profile_MWh = Profiel (MW) × {cfg.MWH_FACTOR}\n"
         f"Hedge_MWh = Totaal inkoop (MW) × {cfg.MWH_FACTOR}\n"
         f"Tekort = MAX(0, Profile_MWh − Hedge_MWh) → inkopen op spotmarkt\n"
         f"Overschot = MAX(0, Hedge_MWh − Profile_MWh) → verkopen op spotmarkt"),

        ("Kosten per interval",
         "Kosten base = Base MW × 0.25 × Base prijs (€/MWh)\n"
         "Kosten peak = Peak MW × 0.25 × Peak prijs (€/MWh)\n"
         "Kosten inkoop totaal = Kosten base + Kosten peak\n"
         "Spot inkoop = Tekort (MWh) × EPEX spotprijs\n"
         "Spot verkoop = Overschot (MWh) × EPEX spotprijs\n"
         "Netto spot = Spot verkoop − Spot inkoop"),

        ("Totale kosten",
         "Totale energiekosten = Kosten inkoop totaal − Netto spotresultaat\n"
         "Kostprijs per MWh = Totale energiekosten / |Totaal profiel|"),

        ("Gebruikte optimalisatie" + (f": {opt['label']}" if opt else ""),
         opt["desc"] if opt else "Onbekend"),
    ]

    row = 3
    for title, body in sections:
        ws.cell(row=row, column=1, value=title).font = _LABEL_FONT
        row += 1
        for line in body.split("\n"):
            ws.cell(row=row, column=1, value=line).font = _DATA_FONT
            row += 1
        row += 1  # lege rij tussen secties

    # Kolomdefinities
    row += 1
    ws.cell(row=row, column=1, value="Kolommen in het Uurdata-tabblad").font = _SECTION_FONT
    row += 1

    ws.cell(row=row, column=1, value="Kolom").font = _HEADER_FONT
    ws.cell(row=row, column=2, value="Beschrijving").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.cell(row=row, column=2).fill = _HEADER_FILL
    row += 1

    col_descriptions = {
        "Datum/tijd": "Tijdstempel van het 15-minuten interval",
        "Kwartaal": "Kwartaal (1-4)",
        "Piekuur (1=ja)": f"1 = piekuur (Ma-Vr {cfg.PEAK_HOUR_START}:00–{cfg.PEAK_HOUR_END}:00), 0 = daluur",
        "Profiel (MW)": "Energieverbruik/-opwek in MW voor dit interval",
        "Base prijs (€/MWh)": "Contractprijs voor het Base-blok",
        "Peak prijs (€/MWh)": "Contractprijs voor het Peak-blok",
        "EPEX spot (€/MWh)": "Day-ahead spotprijs van de EPEX-markt",
        "Base inkoop (MW)": "Ingekocht Base-vermogen (constant)",
        "Peak inkoop (MW)": "Extra Peak-vermogen (alleen tijdens piekuren)",
        "Totaal inkoop (MW)": "= Base + Peak (effectieve inkoop dit uur)",
        "Profiel (MWh)": f"= Profiel (MW) × {cfg.MWH_FACTOR}",
        "Inkoop (MWh)": f"= Totaal inkoop (MW) × {cfg.MWH_FACTOR}",
        "Overschot (MWh)": "= MAX(0, Inkoop − Profiel) — te veel ingekocht",
        "Tekort (MWh)": "= MAX(0, Profiel − Inkoop) — te weinig ingekocht",
        "Benut (MWh)": "= MIN(|Profiel|, |Inkoop|) — effectief gebruikt",
        "Kosten base (€)": f"= Base MW × {cfg.MWH_FACTOR} × Base prijs",
        "Kosten peak (€)": f"= Peak MW × {cfg.MWH_FACTOR} × Peak prijs",
        "Kosten inkoop (€)": "= Kosten base + Kosten peak",
        "Spot inkoop (€)": "= Tekort (MWh) × EPEX spotprijs",
        "Spot verkoop (€)": "= Overschot (MWh) × EPEX spotprijs",
        "Netto spot (€)": "= Spot verkoop − Spot inkoop",
    }

    for _, label, _ in _UURDATA_COLS:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=col_descriptions.get(label, "")).font = _DATA_FONT
        row += 1

    _set_col_widths(ws, {"A": 30, "B": 60})


# ═══════════════════════════════════════════════════════════════════════════
# Publieke API
# ═══════════════════════════════════════════════════════════════════════════

def build_export_workbook(
    df: pd.DataFrame,
    category: str,
    optimization_key: str,
    strategy_period: str,
    epex_loaded: bool,
    q_stats: pd.DataFrame,
) -> io.BytesIO:
    """Bouw een Excel-werkboek met de volledige berekening.

    Parameters
    ----------
    df : DataFrame
        Het DataFrame *na* ``apply_hedge_columns()`` — bevat alle berekende kolommen.
    category : str
        Geselecteerde categorie (Consumer/Prosumer/Producer).
    optimization_key : str
        Key van de geselecteerde optimalisatie (uit config.OPTIMIZATIONS).
    strategy_period : str
        "Per Jaar" of "Per Kwartaal".
    epex_loaded : bool
        Of EPEX spotprijzen beschikbaar zijn.
    q_stats : DataFrame
        Output van ``compute_quarterly_table()``.

    Returns
    -------
    io.BytesIO
        Buffer met het Excel-bestand, klaar voor download.
    """
    wb = Workbook()

    # Haal positie-info uit het DataFrame
    base_mw = df["Hedge_Base_MW"].iloc[0] if "Hedge_Base_MW" in df.columns else 0.0
    peak_add_mw = df["Hedge_Peak_MW"].max() if "Hedge_Peak_MW" in df.columns else 0.0
    price_base = df["Price_Base"].mean() if "Price_Base" in df.columns else 0.0
    price_peak = df["Price_Peak"].mean() if "Price_Peak" in df.columns else 0.0

    # Sheet 1: Scenario (gebruikt wb.active)
    _build_scenario_sheet(
        wb, category, optimization_key,
        strategy_period, base_mw, peak_add_mw, price_base, price_peak,
    )

    # Sheet 4: Uurdata (eerst, zodat we n_rows kennen voor formules)
    n_rows = _build_uurdata_sheet(wb, df)

    # Sheet 2: Berekening (met formules die naar Uurdata verwijzen)
    _build_berekening_sheet(wb, n_rows)

    # Sheet 3: Kwartaaloverzicht
    _build_quarterly_sheet(wb, q_stats)

    # Sheet 5: Uitleg
    _build_uitleg_sheet(wb, optimization_key)

    # Herorden tabs: Scenario, Berekening, Kwartaaloverzicht, Uurdata, Uitleg
    desired_order = ["Scenario", "Berekening", "Kwartaaloverzicht", "Uurdata", "Uitleg"]
    sheet_names = wb.sheetnames
    for i, name in enumerate(desired_order):
        if name in sheet_names:
            current_idx = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=i - current_idx)

    # Opslaan naar buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
